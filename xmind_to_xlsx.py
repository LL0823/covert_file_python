#!/user/bin/env python3
# -*- coding: utf-8 -*-
# author: liu
# time: 2024/3/4 15:11
import os
from typing import List, Any
from xmindparser import xmind_to_dict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment


class ExcelHandle:
    def __init__(self, excel_file_path):
        if not os.path.exists(excel_file_path):
            f = Workbook()
            self.sheet = f.active
            f.save(excel_file_path)
        else:
            f = load_workbook(excel_file_path)
            self.sheet = f.active

    def get_active_sheet(self):
        return self.sheet


def xmind_read(file_name):
    """
    读取xmind成dict格式
    :param file_name:
    :return: 返回去掉画布和主主题的字典
    """
    xmind_json = xmind_to_dict(file_name)
    return xmind_json[0]['topic']['topics']


def dict_to_lists(dict_: dict, lists: list, title: str):
    """
    递归去分解字典，如果字典里的topics中存在，就要走到最里面一层
    :param dict_:
    :param lists:
    :param title:
    :return:
    """
    # 去除title首尾空格
    title = title.strip()
    # 若title为空，则直接取value
    if len(title) == 0:
        concat_title = dict_["title"].strip()
    else:
        concat_title = title + "\t" + dict_["title"].strip()
    if not dict_.__contains__("topics"):
        lists.append(concat_title)
    else:
        for d in dict_["topics"]:
            dict_to_lists(d, lists, concat_title)


def get_xmind_list(xmind_list: dict):
    """
    把xmind读取出来的字典转换成一行一行的数据
    :return:
    """
    lists: List[Any] = []
    for i in range(len(xmind_list)):
        dict_to_lists(xmind_list[i], lists, '')
    for n in range(len(lists)):
        # 把用\n分割的列表元素，分开
        lists[n] = lists[n].split('\t')
    # print(lists)
    return lists


def save_excel(xmind_list, excel_path):
    f = Workbook()
    sheet = f.active
    # 第一行固定的表头标题写入
    row_header = ["序号", "模块", "标题", "步骤", "预期结果", "优先级"]
    for i in range(len(row_header)):
        sheet.cell(1, i + 1, row_header[i])
    # 内容写入
    for x in range(len(xmind_list)):
        sheet.cell(x + 2, 1, x + 1)
        for y in range(len(xmind_list[x])):
            sheet.cell(x + 2, y + 2, xmind_list[x][y])
    f.save(excel_path)


def excel_merge_cells(excel_path):
    """
    根据列的内容进行合并。仅合并前两列
    :param excel_path:
    :return:
    """
    f = load_workbook(excel_path)
    sheet1 = f.active
    rows = sheet1.max_row
    cols = sheet1.max_column
    # 只合并前两列单元格
    count_str = ''
    count = 0  # 计数，记录连续几个重复
    start_row = 1
    for col in range(1, 4):
        for row in range(1, rows + 1):
            if count_str != sheet1.cell(row, col).value:
                if count != 0:
                    # print(f'start_row={start_row}, end_row={start_row + count}, start_column={i}, end_column={i}')
                    # 和前一个字段不一样，并且计数不为0，就把前几个一样的数据进行合并。这里要重构一下
                    sheet1.merge_cells(start_row=start_row, end_row=start_row + count, start_column=col, end_column=col)
                    align = Alignment(horizontal='left', vertical='center')
                    sheet1.cell(start_row, col).alignment = align
                    count = 0
                count_str = sheet1.cell(row, col).value
            else:
                if count == 0:
                    start_row = row - 1
                count += 1
    f.save(excel_path)


def file_name_joint(file_path):
    excel_name = file_path.split('\\')[-1].split(".")[0] + '.xlsx'
    excel_path = os.path.join(os.getcwd(), excel_name)
    return excel_path


def run(xmind_path, is_excel_merge):
    """
    读取xmind为字典格式；对字典进行处理，递归成列表；将列表数据放入excel；excel格式整理
    :param is_excel_merge: 是否需要合并单元格
    :param xmind_path:要转换的xmind文件路径
    :return:
    """
    xmind_json = xmind_read(xmind_path)
    excel_path = file_name_joint(xmind_path)
    xmind_list = get_xmind_list(xmind_json)
    save_excel(xmind_list, excel_path)
    if is_excel_merge:
        excel_merge_cells(excel_path)
    return excel_path

# if __name__ == '__main__':
#     xmind_path_ = r"D:\003_workspace\pythonExec\xmind_to_excel\用例V1.0.xmind"
#     run(xmind_path_)
